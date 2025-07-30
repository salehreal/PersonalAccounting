from PyQt5.QtWidgets import QApplication, QWidget, QComboBox, QInputDialog, QMessageBox, QTableWidgetItem, QVBoxLayout, QLabel, QListWidgetItem, QFileDialog
from PyQt5.uic import loadUi
from PyQt5 import uic, QtWidgets, QtCore
from sms import send_sms
from PyQt5.QtCore import QTimer, QDate, Qt
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QLineSeries, QCategoryAxis, QValueAxis
from PyQt5.QtGui import QPainter, QPixmap, QColor, QFont
from datetime import datetime, timedelta
from dbfunctions import get_user_fullname, get_user_id_by_phone
import sys
import re
import os
import dbfunctions
import random
import subprocess
import jdatetime
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, date
from PyQt5.QtChart import QChart, QChartView, QPieSeries
from dateutil.relativedelta import relativedelta
from dbfunctions import remove_category
from dbfunctions import connect


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
    return os.path.join(base_path, relative_path)

def fa_to_en(text):
    fa_digits = '۰۱۲۳۴۵۶۷۸۹'
    en_digits = '0123456789'
    return text.translate(str.maketrans(fa_digits, en_digits))

def show_messagebox(parent, title, text, icon=QMessageBox.Information):
    msg = QMessageBox(parent)
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setIcon(icon)
    msg.setStyleSheet("""
        QMessageBox {
            background-color: rgb(0, 92, 137); color: white;
            font-size: 15px;
        }
        QPushButton {
            font: 16pt ".AppleSystemUIFont"; background-color:rgb(109, 171, 231); color: white; border-radius: 6px; padding: 6px 12px; font-weight: bold; 
        }
        QPushButton:hover {
            background-color: rgb(109, 160, 200);
        }
    """)
    msg.exec_()

dbfunctions.create_tables()


class Main(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('ui/mainpage.ui'), self)

        self.signinbutton.clicked.connect(self.ShowSignInPage)
        self.signupbutton.clicked.connect(self.ShowSignUpPage)
        
    def ShowSignInPage(self):
        window2.show()
        self.close()

    def ShowSignUpPage(self):
        window3.show()
        self.close()


class SignInPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('ui/signinpage.ui'), self)
        self.backbutton.clicked.connect(self.ShowMainPage)
        self.signinbutton.clicked.connect(self.CheckUser)

    def ShowMainPage(self):
        window1.show()
        self.close()

    def CheckUser(self):
        def fa_to_en(text):
            fa_digits = '۰۱۲۳۴۵۶۷۸۹'
            en_digits = '0123456789'
            return text.translate(str.maketrans(fa_digits, en_digits))

        phone = fa_to_en(self.phonelineedit.text().strip())
        password = fa_to_en(self.passwordlineedit.text().strip())
        alluser = dbfunctions.check_user()

        if phone and password:
            for user in alluser:
                db_phone = user['phone']
                db_password = user['password']
                db_id = user['id']

                if db_phone == phone and db_password == password:
                    self.errorlabel.setText('خوش آمدید')
                    global window5
                    window5 = WorkPage(db_id)
                    window5.show()
                    self.close()
                    return

            self.errorlabel.setText('شماره موبایل یا رمز عبور اشتباه است')
        else:
            self.errorlabel.setText('لطفاً اطلاعات را به صورت کامل وارد کنید')


class SignUpPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('ui/signupage.ui'), self)
        self.backbutton.clicked.connect(self.ShowMainPage)
        self.signupbutton.clicked.connect(self.AddUser)

    def ShowMainPage(self):
        window1.show()
        self.close()

    def AddUser(self):
        def fa_to_en(text):
            fa_digits = '۰۱۲۳۴۵۶۷۸۹'
            en_digits = '0123456789'
            return text.translate(str.maketrans(fa_digits, en_digits))

        fullname = self.fullnamelineedit.text()
        password = fa_to_en(self.passwordlineedit.text())
        repeat = fa_to_en(self.repeatlineedit.text())
        phone = fa_to_en(self.phonelineedit.text())
        alluser = dbfunctions.check_user()

        if all([fullname, password, repeat, phone]):
            if re.fullmatch(r'^(09[0-9]{9})$', phone):
                if len(password) >= 6:
                    if not re.fullmatch(r'[A-Za-z0-9@#$%^&+=!]{6,}', password):
                        self.errorlabel.setText('رمز عبور باید فقط شامل حروف و اعداد انگلیسی باشد')
                        return
                    if password == repeat:
                        for user in alluser:
                            if user['phone'] == phone:
                                self.errorlabel.setText('کاربری با این شماره موبایل از قبل موجود است')
                                return
                            
                        self.pending_user = {
                            'fullname': fullname,
                            'password': password,
                            'phone': phone
                        }
                        window4.set_otp(self)
                        window4.show()
                        self.close()
                    else:
                        self.errorlabel.setText('رمز عبور با تکرار آن مطابقت ندارد')
                else:
                    self.errorlabel.setText('رمز عبور باید حداقل ۶ کاراکتر باشد')
            else:
                self.errorlabel.setText('شماره موبایل معتبر نیست')
        else:
            self.errorlabel.setText('لطفاً همه فیلدها را پر کنید')


class OtpPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('ui/otppage.ui'), self)

        self.generated_code = None
        self.signup_page = None
        self.timer = QTimer()
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self.update_timer)
        self.remaining_seconds = 0

        self.confirmbutton.clicked.connect(self.verify_code)
        self.resendbutton.clicked.connect(self.resend_code)
        self.backbutton.clicked.connect(self.go_back)

    def fa_to_en(self, text):
        fa_digits = '۰۱۲۳۴۵۶۷۸۹'
        en_digits = '0123456789'
        return text.translate(str.maketrans(fa_digits, en_digits))

    def set_otp(self, signup_page):
        self.signup_page = signup_page
        self.send_new_code()
        self.start_resend_timer()

    def send_new_code(self):
        self.generated_code = str(random.randint(1000, 9999))
        self.generated_time = datetime.now()
        phone = self.signup_page.phonelineedit.text()
        success = send_sms(phone, self.generated_code)

        if success:
            self.messageLabel.setText("کد تایید برای شما ارسال شد. لطفا آن را وارد نمایید")
        else:
            self.messageLabel.setText(" ارسال ناموفق بود ")

    def start_resend_timer(self):
        self.resendbutton.setEnabled(False)
        self.remaining_seconds = 120
        self.resendbutton.setText(f"ارسال مجدد ({self.remaining_seconds})")
        self.timer.start()

    def update_timer(self):
        self.remaining_seconds -= 1
        if self.remaining_seconds > 0:
            self.resendbutton.setText(f"ارسال مجدد ({self.remaining_seconds})")
        else:
            self.timer.stop()
            self.resendbutton.setText("ارسال مجدد")
            self.resendbutton.setEnabled(True)

    def resend_code(self):
        self.send_new_code()
        self.start_resend_timer()

    def verify_code(self):
        entered = self.otplineedit.text().strip()
        entered = self.fa_to_en(entered)

        if datetime.now() - self.generated_time > timedelta(minutes=2):
            self.confirmbutton.setText("⏰ کد منقضی شده")
            return

        if entered == self.generated_code:
            user = self.signup_page.pending_user
            from dbfunctions import insert_user, get_user_id_by_phone

            insert_user(user['fullname'], user['password'], user['phone'])
            user_id = get_user_id_by_phone(user['phone'])
            if isinstance(user_id, (list, tuple)):
                user_id = user_id[0]
            try:
                user_id = int(user_id)
            except Exception:
                error_text = (
                    "<b style='color:#c00;'>خطا:</b><br>"
                    "شناسه کاربر نامعتبر است!<br><br>"
                )

                msgbox = QtWidgets.QMessageBox()
                msgbox.setIcon(QtWidgets.QMessageBox.Warning)
                msgbox.setWindowTitle("خطا")
                msgbox.setTextFormat(QtCore.Qt.RichText)
                msgbox.setText(error_text)
                msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msgbox.setDefaultButton(QtWidgets.QMessageBox.Ok)
                msgbox.exec()
                
                return


            self.confirmbutton.setText("تأیید شد")
            global window5
            window5 = WorkPage(user_id)
            window5.show()
            self.close()
        else:
            self.confirmbutton.setText("کد نادرست")

    def go_back(self):
        window3.show()
        self.close()


class WorkPage(QWidget):
    def __init__(self, user_id):
        super().__init__()
        uic.loadUi(resource_path("ui/workpage.ui"), self)

        try:
            user_id = int(user_id)
        except Exception:
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "شناسه کاربر نامعتبر است<br><br>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.setDefaultButton(QtWidgets.QMessageBox.Ok)
            msgbox.exec()

            user_id = None

        fullname = get_user_fullname(user_id) if user_id is not None else "--"
        self.fullnamelabel.setText(f"سلام {fullname} عزیز!")

        self.ConfirmEventButton.clicked.connect(self.ShowIncomePage)
        self.AccountsButton.clicked.connect(self.ShowAccountsPage)
        self.FinancialReportButton.clicked.connect(self.ShowFinancialReportPage)
        self.EventsButton.clicked.connect(self.ShowEventsPage)
        self.CategoriesButton.clicked.connect(self.ShowCategoriesPage)

    def ShowIncomePage(self):
        global income_window
        income_window = AddEventPage()
        income_window.show()

    def ShowAccountsPage(self):
        global account_window
        account_window = AddAccountPage()
        account_window.show()

    def ShowFinancialReportPage(self):
        global report_window
        report_window = FinancialReportPage()
        report_window.show()
    
    def ShowEventsPage(self):
        global report_window
        report_window = EventsPage()
        report_window.show()

    def ShowCategoriesPage(self):
        global categories_window
        categories_window = CategoriesPage()
        categories_window.show()


class AddEventPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('ui/addevent.ui'), self)

        today_jalali = jdatetime.date.today()
        formatted_date = today_jalali.strftime('%Y/%m/%d')
        self.dateLineEdit.setText(formatted_date)
        self.dateLineEdit.setPlaceholderText('مثال: ۱۴۰۴/۰۴/۲۵')

        self.typeComboBox.setEditable(False)
        self.categoryComboBox.setEditable(True)
        self.accountComboBox.setEditable(False)
        self.typeComboBox.addItems(['درآمد', 'هزینه'])

        self.typeComboBox.currentIndexChanged.connect(self.update_category_combo)
        self.ConfirmEventButton.clicked.connect(self.save_event)
        self.backbutton.clicked.connect(self.close)
        self.CostLineEdit.textChanged.connect(self.format_amount)

        self.update_category_combo()
        self.load_accounts()

    def fa_to_en(self, text):
        fa_digits = '۰۱۲۳۴۵۶۷۸۹'
        en_digits = '0123456789'
        return text.translate(str.maketrans(fa_digits, en_digits))

    def is_valid_jalali_date(self, date_str):
        try:
            y, m, d = map(int, date_str.split("/"))
            jdatetime.date(y, m, d)
            return True
        except:
            return False

    def format_amount(self, text):
        raw = self.fa_to_en(text).replace(",", "")
        if raw.isdigit():
            formatted = "{:,}".format(int(raw))
            cursor_pos = self.CostLineEdit.cursorPosition()
            self.CostLineEdit.blockSignals(True)
            self.CostLineEdit.setText(formatted)
            self.CostLineEdit.blockSignals(False)
            delta = len(formatted) - len(raw)
            self.CostLineEdit.setCursorPosition(cursor_pos + delta)

    def update_category_combo(self):
        selected_type = self.typeComboBox.currentText().strip()
        db_type = 'income' if selected_type == 'درآمد' else 'expense'

        self.categoryComboBox.clear()
        conn = dbfunctions.connect()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM categories WHERE type = ?", (db_type,))
        rows = cursor.fetchall()
        conn.close()

        for row in rows:
            self.categoryComboBox.addItem(row[0])

    def load_accounts(self):
        self.accountComboBox.clear()
        conn = dbfunctions.connect()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM accounts")
        rows = cursor.fetchall()
        conn.close()
        for row in rows:
            self.accountComboBox.addItem(row[0])

    def save_event(self):
        amount = self.fa_to_en(self.CostLineEdit.text().replace(",", "").strip())
        type_value = self.typeComboBox.currentText().strip()
        category_name = self.categoryComboBox.currentText().strip()
        account_name = self.accountComboBox.currentText().strip()
        date_text = self.fa_to_en(self.dateLineEdit.text().strip())
        description = self.textEdit.toPlainText().strip()

        if not all([amount, type_value, category_name, account_name, date_text]):
            warning_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "لطفاً همه فیلدهای اجباری را پر کنید<br><br>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.setDefaultButton(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        if not self.is_valid_jalali_date(date_text):
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "فرمت تاریخ نامعتبر است.<br>"
                "<i>مثال: ۱۴۰۴/۰۴/۲۵</i>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        if not amount.isdigit():
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "مبلغ باید فقط شامل <b>عدد</b> باشد."
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        db_type = 'income' if type_value == 'درآمد' else 'expense'

        conn = dbfunctions.connect()
        cursor = conn.cursor()

        cursor.execute("SELECT id FROM categories WHERE name = ? AND type = ?", (category_name, db_type))
        result = cursor.fetchone()
        category_id = result[0] if result else cursor.execute(
            "INSERT INTO categories (name, type) VALUES (?, ?)", (category_name, db_type)
        ).lastrowid

        cursor.execute("SELECT id FROM accounts WHERE name = ?", (account_name,))
        result = cursor.fetchone()
        account_id = result[0] if result else cursor.execute(
            "INSERT INTO accounts (name) VALUES (?)", (account_name,)
        ).lastrowid

        cursor.execute("""
            INSERT INTO transactions (amount, date, category_id, account_id, description)
            VALUES (?, ?, ?, ?, ?)
        """, (amount, date_text, category_id, account_id, description))

        conn.commit()
        conn.close()

        success_text = (
            "<b style='color:green;'>ثبت شد:</b><br>"
            "<span style='color:#333;'>رویداد با موفقیت ثبت شد ✅</span>"
        )

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Information)
        msgbox.setWindowTitle("ثبت موفق")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(success_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msgbox.exec()
        self.close()


class AddAccountPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path("ui/addaccount.ui"), self)

        self.addButton.clicked.connect(self.add_account)
        self.deleteButton.clicked.connect(self.delete_account)
        self.backbutton.clicked.connect(self.close)

        self.load_accounts()

    def load_accounts(self):
        self.accountListWidget.clear()
        conn = dbfunctions.connect()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM accounts")
        rows = cursor.fetchall()
        conn.close()

        for row in rows:
            item = QListWidgetItem(row[0])
            item.setTextAlignment(Qt.AlignRight)
            self.accountListWidget.addItem(item)

    def add_account(self):
        name = self.accountLineEdit.text().strip()
        if not name:
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "لطفاً <b>نام حساب</b> را وارد کنید."
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return


        conn = dbfunctions.connect()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM accounts WHERE name = ?", (name,))
        result = cursor.fetchone()

        if result:
            info_text = (
                "<b style='color:#f57c00;'>هشدار:</b><br>"
                "<span style='color:#333;'>این حساب قبلاً ثبت شده است </span>"
            )
        else:
            cursor.execute("INSERT INTO accounts (name) VALUES (?)", (name,))
            conn.commit()
            info_text = (
                "<b style='color:green;'>ثبت شد:</b><br>"
                "<span style='color:#333;'>حساب جدید با موفقیت اضافه شد </span>"
            )
            self.accountLineEdit.clear()
            self.load_accounts()

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Information)
        msgbox.setWindowTitle("اطلاع رسانی")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(info_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msgbox.exec()

        conn.close()

    def delete_account(self):
        selected_item = self.accountListWidget.currentItem()
        if not selected_item:
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "لطفاً یک <b>حساب</b> را برای حذف انتخاب کنید "
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        account_name = selected_item.text()

        conn = dbfunctions.connect()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM transactions
            WHERE account_id IN (
                SELECT id FROM accounts WHERE name = ?
            )
        """, (account_name,))
        count = cursor.fetchone()[0]

        if count > 0:
            error_text = (
                "<b style='color:#c00;'>امکان حذف نیست:</b><br>"
                "<span style='color:#333;'>این حساب در <b>تراکنش‌ها</b> استفاده شده و قابل حذف نیست </span>"
            )
        else:
            cursor.execute("DELETE FROM accounts WHERE name = ?", (account_name,))
            conn.commit()
            error_text = (
                "<b style='color:green;'>حذف شد:</b><br>"
                "<span style='color:#333;'>حساب با موفقیت حذف شد </span>"
            )
            self.load_accounts()

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Information)
        msgbox.setWindowTitle("وضعیت حذف")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(error_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msgbox.exec()

        conn.close()


class FinancialReportPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path('ui/financialreport.ui'), self)

        self.setLayoutDirection(Qt.RightToLeft)
        self.fromLineEdit.setPlaceholderText("مثال: ۱۴۰۴/۰۴/۰۱")
        self.toLineEdit.setPlaceholderText("مثال: ۱۴۰۴/۰۴/۳۰")
        self.yearLineEdit.setPlaceholderText("مثال: ۱۴۰۴")

        self.exportToExcelButton.clicked.connect(self.export_to_excel)
        self.generateReportButton.clicked.connect(self.generate_report)
        self.backButton.clicked.connect(self.close)
        self.exportToExcelYearlyButton.clicked.connect(self.export_to_excel_yearly)
        self.generateYearlyReportButton.clicked.connect(self.generate_yearly_report)

        self.TrendChartLayout = QVBoxLayout()
        self.TrendChartContainer.setLayout(self.TrendChartLayout)
        self.TrendChartContainer.setStyleSheet("background-color: #eefaff;")

        self.expenseChartLayout = QVBoxLayout()
        self.expenseChartContainer.setLayout(self.expenseChartLayout)

        self.incomeChartLayout = QVBoxLayout()
        self.incomeChartContainer.setLayout(self.incomeChartLayout)

    def fa_to_en(self, text):
        fa_digits = '۰۱۲۳۴۵۶۷۸۹'
        en_digits = '0123456789'
        return text.translate(str.maketrans(fa_digits, en_digits))

    def is_valid_jalali_date(self, date_str):
        try:
            y, m, d = map(int, date_str.split("/"))
            jdatetime.date(y, m, d)
            return True
        except:
            return False

    def convert_type_to_farsi(self, type_en):
        return "درآمد" if type_en == "income" else "هزینه"

    def generate_report(self):
        from_date = self.fa_to_en(self.fromLineEdit.text().strip())
        to_date = self.fa_to_en(self.toLineEdit.text().strip())

        if not from_date or not to_date:
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "لطفاً <b>بازه زمانی</b> را کامل وارد کنید "
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        if not self.is_valid_jalali_date(from_date) or not self.is_valid_jalali_date(to_date):
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "تاریخ واردشده معتبر نیست.<br>"
                "<i>لطفاً مانند ۱۴۰۴/۰۴/۲۵ وارد کنید.</i> 📅"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        conn = dbfunctions.connect()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT SUM(t.amount)
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE c.type = 'income' AND t.date BETWEEN ? AND ?
        """, (from_date, to_date))
        income = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT SUM(t.amount)
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE c.type = 'expense' AND t.date BETWEEN ? AND ?
        """, (from_date, to_date))
        expense = cursor.fetchone()[0] or 0

        self.totalIncomeLabel.setText(f"{income:,} ریال")
        self.totalExpenseLabel.setText(f"{expense:,} ریال")
        self.netBalanceLabel.setText(f"{(income - expense):,} ریال")

        cursor.execute("""
            SELECT c.name, c.type, t.amount, t.date
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE t.date BETWEEN ? AND ?
            ORDER BY t.date ASC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        self.categoryTable.setRowCount(len(rows))
        self.categoryTable.setColumnCount(4)
        self.categoryTable.setHorizontalHeaderLabels(["دسته", "نوع", "مبلغ", "تاریخ"])

        for i, (cat, typ, amt, date_str) in enumerate(rows):
            type_fa = self.convert_type_to_farsi(typ)
            amount_text = f"{amt:,} ریال"
            if typ == "expense":
                amount_text = f"({amount_text})"

            self.categoryTable.setItem(i, 0, QTableWidgetItem(cat))
            self.categoryTable.setItem(i, 1, QTableWidgetItem(type_fa))
            self.categoryTable.setItem(i, 2, QTableWidgetItem(amount_text))
            self.categoryTable.setItem(i, 3, QTableWidgetItem(date_str))
            for j in range(4):
                self.categoryTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        cursor.execute("""
            SELECT a.name,
                SUM(CASE WHEN c.type = 'expense' THEN -t.amount ELSE t.amount END)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.id
            JOIN categories c ON t.category_id = c.id
            WHERE t.date BETWEEN ? AND ?
            GROUP BY a.name
        """, (from_date, to_date))
        acc_rows = cursor.fetchall()
        self.accountTable.setRowCount(len(acc_rows))
        self.accountTable.setColumnCount(2)
        self.accountTable.setHorizontalHeaderLabels(["حساب", "موجودی"])

        cursor.execute("""
            SELECT c.name, SUM(t.amount) AS total_income
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE c.type = 'income' AND t.date BETWEEN ? AND ?
            GROUP BY c.name
            ORDER BY total_income DESC
        """, (from_date, to_date))
        income_rows = cursor.fetchall()
        self.categoryincomeTable.setRowCount(len(income_rows))
        self.categoryincomeTable.setColumnCount(2)
        self.categoryincomeTable.setHorizontalHeaderLabels(["دسته‌بندی", "مجموع درآمد"])

        for i, (cat_name, total) in enumerate(income_rows):
            total_text = f"{total:,} ریال" if total else "۰ ریال"
            self.categoryincomeTable.setItem(i, 0, QTableWidgetItem(cat_name))
            self.categoryincomeTable.setItem(i, 1, QTableWidgetItem(total_text))
            for j in range(2):
                self.categoryincomeTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        for i, (acc, amt) in enumerate(acc_rows):
            amount_text = f"{abs(amt):,} ریال"
            if amt < 0:
                amount_text = f"({amount_text})"
            self.accountTable.setItem(i, 0, QTableWidgetItem(acc))
            self.accountTable.setItem(i, 1, QTableWidgetItem(amount_text))
            for j in range(2):
                self.accountTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        cursor.execute("""
            SELECT c.name, SUM(t.amount) AS total_expense
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE c.type = 'expense' AND t.date BETWEEN ? AND ?
            GROUP BY c.name
            ORDER BY total_expense DESC
        """, (from_date, to_date))
        category_rows = cursor.fetchall()
        self.categorycostTable.setRowCount(len(category_rows))
        self.categorycostTable.setColumnCount(2)
        self.categorycostTable.setHorizontalHeaderLabels(["دسته‌بندی", "مجموع هزینه"])

        for i, (cat_name, total) in enumerate(category_rows):
            total_text = f"{total:,} ریال" if total else "۰ ریال"
            self.categorycostTable.setItem(i, 0, QTableWidgetItem(cat_name))
            self.categorycostTable.setItem(i, 1, QTableWidgetItem(total_text))
            for j in range(2):
                self.categorycostTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        conn.close()
        self.show_expense_chart()
        self.show_income_chart()
        self.show_trend_chart()

    def generate_yearly_report(self):
        year_raw = self.yearLineEdit.text().strip()
        year = self.fa_to_en(year_raw)

        if not year or not year.isdigit():
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "لطفاً یک <b>سال معتبر</b> وارد کنید"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        like_pattern = f"{year}/%"
        from_date = f"{year}/01/01"
        to_date = f"{year}/12/29"

        conn = dbfunctions.connect()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT c.name, c.type, t.amount, t.date
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE t.date LIKE ?
            ORDER BY t.date ASC
        """, (like_pattern,))
        rows = cursor.fetchall()

        self.categoryTable.setRowCount(len(rows))
        self.categoryTable.setColumnCount(4)
        self.categoryTable.setHorizontalHeaderLabels(["دسته", "نوع", "مبلغ", "تاریخ"])

        total_income, total_expense = 0, 0

        for i, (cat, typ, amt, date_str) in enumerate(rows):
            type_fa = self.convert_type_to_farsi(typ)
            amount_text = f"{amt:,} ریال"
            if typ == "expense":
                amount_text = f"({amount_text})"
                total_expense += amt
            else:
                total_income += amt

            self.categoryTable.setItem(i, 0, QTableWidgetItem(cat))
            self.categoryTable.setItem(i, 1, QTableWidgetItem(type_fa))
            self.categoryTable.setItem(i, 2, QTableWidgetItem(amount_text))
            self.categoryTable.setItem(i, 3, QTableWidgetItem(date_str))
            for j in range(4):
                self.categoryTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        self.totalIncomeLabel.setText(f"{total_income:,} ریال")
        self.totalExpenseLabel.setText(f"{total_expense:,} ریال")
        self.netBalanceLabel.setText(f"{(total_income - total_expense):,} ریال")

        cursor.execute("""
            SELECT a.name,
                SUM(CASE WHEN c.type = 'expense' THEN -t.amount ELSE t.amount END)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.id
            JOIN categories c ON t.category_id = c.id
            WHERE t.date LIKE ?
            GROUP BY a.name
        """, (like_pattern,))
        acc_rows = cursor.fetchall()
        self.accountTable.setRowCount(len(acc_rows))
        self.accountTable.setColumnCount(2)
        self.accountTable.setHorizontalHeaderLabels(["حساب", "موجودی"])

        for i, (acc, amt) in enumerate(acc_rows):
            amount_text = f"{abs(amt):,} ریال"
            if amt < 0:
                amount_text = f"({amount_text})"
            self.accountTable.setItem(i, 0, QTableWidgetItem(acc))
            self.accountTable.setItem(i, 1, QTableWidgetItem(amount_text))
            for j in range(2):
                self.accountTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        cursor.execute("""
            SELECT c.name, SUM(t.amount) AS total_income
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE c.type = 'income' AND t.date BETWEEN ? AND ?
            GROUP BY c.name
            ORDER BY total_income DESC
        """, (from_date, to_date))
        income_rows = cursor.fetchall()
        self.categoryincomeTable.setRowCount(len(income_rows))
        self.categoryincomeTable.setColumnCount(2)
        self.categoryincomeTable.setHorizontalHeaderLabels(["دسته‌بندی", "مجموع درآمد"])

        for i, (cat_name, total) in enumerate(income_rows):
            total_text = f"{total:,} ریال" if total else "۰ ریال"
            self.categoryincomeTable.setItem(i, 0, QTableWidgetItem(cat_name))
            self.categoryincomeTable.setItem(i, 1, QTableWidgetItem(total_text))
            for j in range(2):
                self.categoryincomeTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        cursor.execute("""
            SELECT c.name, SUM(t.amount) AS total_expense
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE c.type = 'expense' AND t.date LIKE ?
            GROUP BY c.name
            ORDER BY total_expense DESC
        """, (like_pattern,))
        cat_rows = cursor.fetchall()
        self.categorycostTable.setRowCount(len(cat_rows))
        self.categorycostTable.setColumnCount(2)
        self.categorycostTable.setHorizontalHeaderLabels(["دسته‌بندی", "مجموع هزینه"])

        for i, (cat_name, total) in enumerate(cat_rows):
            total_text = f"{total:,} ریال" if total else "۰ ریال"
            self.categorycostTable.setItem(i, 0, QTableWidgetItem(cat_name))
            self.categorycostTable.setItem(i, 1, QTableWidgetItem(total_text))
            for j in range(2):
                self.categorycostTable.item(i, j).setTextAlignment(Qt.AlignCenter)

        conn.close()

        self.show_expense_chart()
        self.show_income_chart()
        self.show_trend_chart()

    def show_expense_chart(self):
        while self.expenseChartLayout.count():
            child = self.expenseChartLayout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        data = {}
        for i in range(self.categoryTable.rowCount()):
            if self.categoryTable.item(i, 1).text() == "هزینه":
                cat = self.categoryTable.item(i, 0).text()
                amt_text = self.categoryTable.item(i, 2).text()
                amt_text = amt_text.replace("ریال", "").replace(",", "").replace("(", "").replace(")", "").strip()
                try:
                    amt = int(amt_text)
                    data[cat] = data.get(cat, 0) + amt
                except:
                    pass

        if not data:
            label = QLabel("هیچ هزینه‌ای برای نمایش در نمودار نیست")
            label.setAlignment(Qt.AlignCenter)
            self.expenseChartLayout.addWidget(label)
            return

        series = QPieSeries()
        for cat, amt in data.items():
            series.append(cat, amt)

        chart = QChart()
        chart.addSeries(series)
        chart.legend().setAlignment(Qt.AlignRight)

        chart_view = QChartView(chart)
        chart_view.setRenderHint(QPainter.Antialiasing)
        chart_view.setMinimumHeight(150)
        chart_view.setStyleSheet("border: none;")
        self.expenseChartLayout.addWidget(chart_view)

    def show_income_chart(self):
        while self.incomeChartLayout.count():
            child = self.incomeChartLayout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        data = {}
        for i in range(self.categoryTable.rowCount()):
            if self.categoryTable.item(i, 1).text() == "درآمد":
                cat = self.categoryTable.item(i, 0).text()
                amt_text = self.categoryTable.item(i, 2).text()
                amt_text = amt_text.replace("ریال", "").replace(",", "").replace("(", "").replace(")", "").strip()
                try:
                    amt = int(amt_text)
                    data[cat] = data.get(cat, 0) + amt
                except:
                    pass

        if not data:
            label = QLabel("هیچ درآمدی برای نمایش در نمودار نیست")
            label.setAlignment(Qt.AlignCenter)
            self.incomeChartLayout.addWidget(label)
            return

        series = QPieSeries()
        for cat, amt in data.items():
            series.append(cat, amt)

        chart = QChart()
        chart.addSeries(series)
        chart.legend().setAlignment(Qt.AlignRight)

        chart_view = QChartView(chart)
        chart_view.setRenderHint(QPainter.Antialiasing)
        chart_view.setMinimumHeight(150)
        self.incomeChartLayout.addWidget(chart_view)
        chart_view.setStyleSheet("border: none;")

    def parse_year_to_range(self, year_text):
        try:
            year_int = int(year_text)
            from_date = datetime.strptime(f"{year_int}/01/01", "%Y/%m/%d")
            to_date = datetime.strptime(f"{year_int}/12/29", "%Y/%m/%d")
            return from_date, to_date
        except:
            return None, None

    def show_trend_chart(self):
        while self.TrendChartLayout.count():
            child = self.TrendChartLayout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        monthly_income = {}
        monthly_expense = {}

        year_text = self.yearLineEdit.text().strip()
        from_text = self.fromLineEdit.text().strip()
        to_text = self.toLineEdit.text().strip()

        if year_text:
            from_date, to_date = self.parse_year_to_range(year_text)
        elif from_text and to_text:
            try:
                from_date = datetime.strptime(from_text, "%Y/%m/%d")
                to_date = datetime.strptime(to_text, "%Y/%m/%d")
            except:
                label = QLabel("فرمت تاریخ وارد شده نامعتبر است")
                label.setAlignment(Qt.AlignCenter)
                label.setStyleSheet("color: #c00; font-size: 14px; padding: 10px;")
                self.TrendChartLayout.addWidget(label)
                return
        else:
            label = QLabel("لطفاً یا سال یا بازهٔ زمانی را وارد کنید")
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("color: #555; font-size: 14px; padding: 10px;")
            self.TrendChartLayout.addWidget(label)
            return

        labels = []
        current = from_date.replace(day=1)
        while current <= to_date:
            label = f"{str(current.year)[-2:]}/{current.month:02d}"
            labels.append(label)
            current += relativedelta(months=1)

        for i in range(self.categoryTable.rowCount()):
            type_item = self.categoryTable.item(i, 1)
            amount_item = self.categoryTable.item(i, 2)
            date_item = self.categoryTable.item(i, 3)

            if not type_item or not amount_item or not date_item:
                continue

            date_text = date_item.text().strip()
            try:
                transaction_date = datetime.strptime(date_text, "%Y/%m/%d")
            except:
                continue

            if transaction_date < from_date or transaction_date > to_date:
                continue

            key = f"{str(transaction_date.year)[-2:]}/{transaction_date.month:02d}"
            type_fa = type_item.text().strip()
            amount_text = amount_item.text().replace("ریال", "").replace(",", "").replace("(", "").replace(")", "").strip()
            try:
                amount = int(amount_text)
            except:
                continue

            if type_fa == "درآمد":
                monthly_income[key] = monthly_income.get(key, 0) + amount
            elif type_fa == "هزینه":
                monthly_expense[key] = monthly_expense.get(key, 0) + amount

        total = sum(monthly_income.values()) + sum(monthly_expense.values())
        if total == 0:
            label = QLabel("هیچ درآمد یا هزینه‌ای برای نمایش در نمودار نیست")
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("color: #555; font-size: 16px; padding: 20px;")
            self.TrendChartLayout.addWidget(label)
            return

        income_series = QLineSeries()
        income_series.setName("درآمد ماهانه")

        expense_series = QLineSeries()
        expense_series.setName("هزینه ماهانه")

        for idx, label in enumerate(labels):
            income_val = monthly_income.get(label, 0)
            expense_val = monthly_expense.get(label, 0)
            income_series.append(idx, income_val)
            expense_series.append(idx, expense_val)

        chart = QChart()
        chart.addSeries(income_series)
        chart.addSeries(expense_series)
        chart.legend().setAlignment(Qt.AlignBottom)

        axisX = QCategoryAxis()
        axisX.setLabelsPosition(QCategoryAxis.AxisLabelsPositionOnValue)
        font = QFont()
        font.setPointSize(8)
        axisX.setLabelsFont(font)
        for idx, label in enumerate(labels):
            axisX.append(label, idx)
        chart.addAxis(axisX, Qt.AlignBottom)
        income_series.attachAxis(axisX)
        expense_series.attachAxis(axisX)

        axisY = QValueAxis()
        axisY.setLabelFormat("%d")
        axisY.setTitleText("مبلغ (ریال)")
        chart.addAxis(axisY, Qt.AlignLeft)
        income_series.attachAxis(axisY)
        expense_series.attachAxis(axisY)

        chart_view = QChartView(chart)
        chart_view.setRenderHint(QPainter.Antialiasing)
        chart_view.setMinimumHeight(200)
        chart_view.setStyleSheet("border: none;")
        chart.setBackgroundBrush(QColor("#ffffff"))

        self.TrendChartLayout.addWidget(chart_view)

    def export_to_excel(self):
        try:
            if self.categoryTable.rowCount() == 0:
                warning_text = (
                    "<b style='color:#f57c00;'>هشدار:</b><br>"
                    "<span style='color:#333;'>ابتدا <b>گزارش‌گیری</b> کنید تا داده‌ها برای خروجی آماده شوند.</span> "
                )

                msgbox = QtWidgets.QMessageBox()
                msgbox.setIcon(QtWidgets.QMessageBox.Warning)
                msgbox.setWindowTitle("هشدار")
                msgbox.setTextFormat(QtCore.Qt.RichText)
                msgbox.setText(warning_text)
                msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msgbox.exec()
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش دسته بندی"
            ws.sheet_view.rightToLeft = True

            bnazanin_font = Font(name="BNazanin", size=12)

            header = ["دسته", "نوع", "مبلغ", "تاریخ"]
            ws.append(header)
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="center")
                cell.font = bnazanin_font

            for row in range(self.categoryTable.rowCount()):
                cat = self.categoryTable.item(row, 0).text()
                typ = self.categoryTable.item(row, 1).text()
                amt_text = self.categoryTable.item(row, 2).text()
                date_str = self.categoryTable.item(row, 3).text()

                amt_clean = amt_text.replace("ریال", "").replace(",", "").replace("(", "").replace(")", "").strip()
                try:
                    amount_value = int(amt_clean)
                    if typ == "هزینه":
                        amount_value = -abs(amount_value)
                    else:
                        amount_value = abs(amount_value)
                except:
                    amount_value = amt_clean

                ws.append([cat, typ, amount_value, date_str])

            for data_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in data_row:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = bnazanin_font

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 3

            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            from_date_raw = self.fromLineEdit.text().strip()
            to_date_raw = self.toLineEdit.text().strip()
            from_date = self.fa_to_en(from_date_raw).replace("/", "-")
            to_date = self.fa_to_en(to_date_raw).replace("/", "-")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"گزارش مالی_{from_date}_تا_{to_date}_{timestamp}.xlsx"
            full_path = os.path.join(desktop_path, filename)

            wb.save(full_path)

            success_text = (
                "<b style='color:green;'>موفقیت:</b><br>"
                "فایل اکسل با موفقیت روی <b>دسکتاپ</b> ذخیره شد. <br><br>"
                f"<span style='color:#555;'>{full_path}</span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Information)
            msgbox.setWindowTitle("ذخیره‌سازی موفق")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(success_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()

            if os.name == "posix":
                subprocess.call(["open", full_path])
            elif os.name == "nt":
                os.startfile(full_path)

        except Exception as e:
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "خطا در <b>ذخیره فایل اکسل</b> رخ داد <br><br>"
                f"<span style='color:#555;'>{str(e)}</span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Critical)
            msgbox.setWindowTitle("خطای سیستمی")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()

    def export_to_excel_yearly(self):
        try:
            if self.categoryTable.rowCount() == 0:
                warning_text = (
                    "<b style='color:#f57c00;'>هشدار:</b><br>"
                    "<span style='color:#333;'>ابتدا <b>گزارش سالیانه</b> را تولید کنید. </span>"
                )

                msgbox = QtWidgets.QMessageBox()
                msgbox.setIcon(QtWidgets.QMessageBox.Warning)
                msgbox.setWindowTitle("هشدار")
                msgbox.setTextFormat(QtCore.Qt.RichText)
                msgbox.setText(warning_text)
                msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msgbox.exec()
                return

            monthly_data = {}
            for i in range(self.categoryTable.rowCount()):
                date_str = self.categoryTable.item(i, 3).text()
                month = date_str.split("/")[1]
                typ = self.categoryTable.item(i, 1).text()
                amount_text = self.categoryTable.item(i, 2).text().replace("ریال", "").replace(",", "").replace("(", "").replace(")", "").strip()

                try:
                    amt = int(amount_text)
                except:
                    amt = 0

                if month not in monthly_data:
                    monthly_data[month] = {"income": 0, "expense": 0}
                if typ == "درآمد":
                    monthly_data[month]["income"] += amt
                elif typ == "هزینه":
                    monthly_data[month]["expense"] += amt

            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش سالیانه"
            ws.sheet_view.rightToLeft = True
            bnazanin_font = Font(name="BNazanin", size=12)

            ws.append(["ماه", "درآمد", "هزینه", "مانده"])
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="center")
                cell.font = bnazanin_font

            for month in sorted(monthly_data.keys(), key=lambda x: int(x)):
                income = monthly_data[month]["income"]
                expense = monthly_data[month]["expense"]
                balance = income - expense
                ws.append([f"ماه {int(month)}", income, expense, balance])

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = bnazanin_font

            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = max_length + 3

            year = self.fa_to_en(self.yearLineEdit.text().strip())
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"خلاصه سالیانه_{year}_{timestamp}.xlsx"
            full_path = os.path.join(desktop_path, filename)
            wb.save(full_path)

            success_text = (
                "<b style='color:green;'>موفقیت:</b><br>"
                "فایل اکسل با موفقیت روی <b>دسکتاپ</b> ذخیره شد <br><br>"
                f"<span style='color:#555;'>{full_path}</span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Information)
            msgbox.setWindowTitle("ذخیره‌سازی موفق")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(success_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()

            if os.name == "posix":
                subprocess.call(["open", full_path])
            elif os.name == "nt":
                os.startfile(full_path)

        except Exception as e:
            error_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "در <b>ذخیره‌سازی فایل اکسل</b> مشکلی پیش آمد <br><br>"
                f"<span style='color:#555;'>{str(e)}</span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Critical)
            msgbox.setWindowTitle("خطای سیستم")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(error_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()


class EventsPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path("ui/events.ui"), self)

        self.conn = sqlite3.connect("accounting.db")
        self.cursor = self.conn.cursor()

        self.eventsTable.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.eventsTable.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)


        self.eventsTable.setEditTriggers(
            QtWidgets.QAbstractItemView.DoubleClicked |
            QtWidgets.QAbstractItemView.SelectedClicked
        )

        self.typeComboBox.setEditable(False)
        self.typeComboBox.addItems(["همه", "درآمد", "هزینه"])
        self.typeComboBox.currentIndexChanged.connect(self.update_category_combo)
        self.CostLineEdit.textChanged.connect(self.format_amount)
        self.searchButton.clicked.connect(self.search_events)
        self.editButton.clicked.connect(self.save_changes)
        self.removeButton.clicked.connect(self.remove_selected_event)
        self.backbutton.clicked.connect(self.close)

        self.update_category_combo()
        self.load_accounts()
        self.load_events()

    def fa_to_en(self, text):
        fa_digits = '۰۱۲۳۴۵۶۷۸۹'
        en_digits = '0123456789'
        return text.translate(str.maketrans(fa_digits, en_digits))

    def format_amount(self, text):
        raw = self.fa_to_en(text).replace(",", "")
        if raw.isdigit():
            formatted = "{:,}".format(int(raw))
            cursor_pos = self.CostLineEdit.cursorPosition()
            self.CostLineEdit.blockSignals(True)
            self.CostLineEdit.setText(formatted)
            self.CostLineEdit.blockSignals(False)
            delta = len(formatted) - len(raw)
            self.CostLineEdit.setCursorPosition(cursor_pos + delta)

    def update_category_combo(self):
        selected_type = self.typeComboBox.currentText().strip()
        self.categoryComboBox.clear()
        self.categoryComboBox.addItem("همه")

        if selected_type == "همه":
            self.cursor.execute("SELECT name FROM categories")
        else:
            db_type = 'income' if selected_type == "درآمد" else "expense"
            self.cursor.execute("SELECT name FROM categories WHERE type = ?", (db_type,))
        rows = self.cursor.fetchall()
        for row in rows:
            self.categoryComboBox.addItem(row[0])

    def load_accounts(self):
        self.accountComboBox.clear()
        self.accountComboBox.addItem("همه")

        self.cursor.execute("SELECT name FROM accounts")
        rows = self.cursor.fetchall()
        for row in rows:
            self.accountComboBox.addItem(row[0])

    def load_events(self):
        self.cursor.execute("""
            SELECT transactions.id, date,
                   (SELECT name FROM categories WHERE id = category_id),
                   amount,
                   (SELECT name FROM accounts WHERE id = account_id),
                   description
            FROM transactions
        """)
        rows = self.cursor.fetchall()
        self.populate_table(rows)

    def search_events(self):
        query = """
            SELECT transactions.id, date,
                   (SELECT name FROM categories WHERE id = category_id),
                   amount,
                   (SELECT name FROM accounts WHERE id = account_id),
                   description
            FROM transactions
            WHERE 1=1
        """
        params = []

        if self.fromLineEdit.text():
            query += " AND date >= ?"
            params.append(self.fa_to_en(self.fromLineEdit.text()))

        if self.toLineEdit.text():
            query += " AND date <= ?"
            params.append(self.fa_to_en(self.toLineEdit.text()))

        selected_type = self.typeComboBox.currentText().strip()
        if selected_type != "همه":
            db_type = 'income' if selected_type == "درآمد" else "expense"
            query += " AND category_id IN (SELECT id FROM categories WHERE type = ?)"
            params.append(db_type)

        selected_category = self.categoryComboBox.currentText().strip()
        if selected_category != "همه":
            query += " AND category_id IN (SELECT id FROM categories WHERE name = ?)"
            params.append(selected_category)

        selected_account = self.accountComboBox.currentText().strip()
        if selected_account != "همه":
            query += " AND account_id IN (SELECT id FROM accounts WHERE name = ?)"
            params.append(selected_account)

        if self.CostLineEdit.text():
            raw_amount = self.fa_to_en(self.CostLineEdit.text().replace(",", ""))
            query += " AND amount = ?"
            params.append(raw_amount)

        self.cursor.execute(query, params)
        rows = self.cursor.fetchall()
        self.populate_table(rows)

    def load_category_types(self):
        self.cursor.execute("SELECT name, type FROM categories")
        rows = self.cursor.fetchall()
        return {name: type_ for name, type_ in rows}

    def populate_table(self, data):
        category_types = self.load_category_types()

        self.eventsTable.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.eventsTable.setColumnCount(7)
        self.eventsTable.setHorizontalHeaderLabels([
            "شناسه", "تاریخ", "نوع رویداد", "دسته‌بندی", "مبلغ", "حساب", "توضیحات"
        ])
        self.eventsTable.setRowCount(len(data))

        for i, row in enumerate(data):
            if len(row) < 6:
                continue

            category_name = row[2]
            category_type = category_types.get(category_name, "—")
            event_type = "درآمد" if category_type == "income" else "هزینه" if category_type == "expense" else "—"

            reordered_row = [
                row[0],
                row[1],
                event_type,
                category_name,
                row[3],
                row[4],
                row[5],
            ]

            for j, item in enumerate(reordered_row):
                value = str(item)
                if j == 4:
                    try:
                        value = "{:,}".format(int(str(row[3]).replace(",", "")))
                    except:
                        pass

                cell = QtWidgets.QTableWidgetItem(value)
                cell.setTextAlignment(QtCore.Qt.AlignCenter)

                if j == 0:
                    cell.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)

                self.eventsTable.setItem(i, j, cell)

    def save_changes(self):
        for row in range(self.eventsTable.rowCount()):
            event_id = self.eventsTable.item(row, 0).text()
            date = self.fa_to_en(self.eventsTable.item(row, 1).text())
            category = self.eventsTable.item(row, 2).text()
            amount = self.fa_to_en(self.eventsTable.item(row, 3).text().replace(",", ""))
            account = self.eventsTable.item(row, 4).text()
            description = self.eventsTable.item(row, 5).text()

            self.cursor.execute("SELECT id FROM categories WHERE name = ?", (category,))
            category_id = self.cursor.fetchone()
            self.cursor.execute("SELECT id FROM accounts WHERE name = ?", (account,))
            account_id = self.cursor.fetchone()

            if category_id and account_id:
                self.cursor.execute("""
                    UPDATE transactions
                    SET date=?, category_id=?, amount=?, account_id=?, description=?
                    WHERE id=?
                """, (date, category_id[0], amount, account_id[0], description, event_id))

        self.conn.commit()
        success_text = (
            "<b style='color:green;'>ذخیره تغییرات:</b><br>"
            "<span style='color:#333;'>تغییرات با موفقیت ذخیره شد </span>"
        )

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Information)
        msgbox.setWindowTitle("ذخیره موفق")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(success_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msgbox.exec()

    def remove_selected_event(self):
        row = self.eventsTable.currentRow()
        print("Current row is:", self.eventsTable.currentRow())
        if row == -1:
            warning_text = (
                "<b style='color:#f57c00;'>هشدار:</b><br>"
                "<span style='color:#333;'>لطفاً یک <b>رویداد</b> را انتخاب کنید </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("هشدار")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        id_item = self.eventsTable.item(row, 0)
        if id_item is None:
            warning_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "<span style='color:#333;'>شناسهٔ <b>رویداد</b> یافت نشد </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        event_id = id_item.text().strip()
        if not event_id:
            warning_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "<span style='color:#333;'>شناسهٔ <b>رویداد</b> نامعتبر است </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        confirm = QtWidgets.QMessageBox.question(
            self, "حذف رویداد",
            "آیا مطمئن هستید که می‌خواهید این رویداد حذف شود؟",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )

        if confirm == QtWidgets.QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM transactions WHERE id = ?", (event_id,))
                self.conn.commit()
                self.eventsTable.removeRow(row)

                success_text = (
                    "<b style='color:green;'>موفقیت:</b><br>"
                    "<span style='color:#333;'>رویداد با موفقیت حذف شد </span>"
                )

                msgbox = QtWidgets.QMessageBox()
                msgbox.setIcon(QtWidgets.QMessageBox.Information)
                msgbox.setWindowTitle("حذف موفق")
                msgbox.setTextFormat(QtCore.Qt.RichText)
                msgbox.setText(success_text)
                msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msgbox.exec()

            except Exception as e:
                error_text = (
                    "<b style='color:#c00;'>خطا:</b><br>"
                    "در <b>حذف رویداد</b> مشکلی پیش آمد <br><br>"
                    f"<span style='color:#555;'>{str(e)}</span>"
                )

                msgbox = QtWidgets.QMessageBox()
                msgbox.setIcon(QtWidgets.QMessageBox.Critical)
                msgbox.setWindowTitle("خطای سیستم")
                msgbox.setTextFormat(QtCore.Qt.RichText)
                msgbox.setText(error_text)
                msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msgbox.exec()


class CategoriesPage(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi(resource_path("ui/categories.ui"), self)

        self.conn = connect()
        self.cursor = self.conn.cursor()
        self.cursor.execute("PRAGMA foreign_keys = ON")

        self.categoriesTable.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.categoriesTable.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.categoriesTable.setEditTriggers(
            QtWidgets.QAbstractItemView.DoubleClicked |
            QtWidgets.QAbstractItemView.SelectedClicked
        )

        self.typeComboBox.addItems(["درآمد", "هزینه"])
        self.typeComboBox.currentIndexChanged.connect(self.update_category_table)
        self.categoryComboBox.currentTextChanged.connect(self.update_edit_button_state)
        self.categoryComboBox.setEditable(True)

        self.addButton.clicked.connect(self.add_category)
        self.editButton.clicked.connect(self.edit_category)
        self.removeButton.clicked.connect(self.remove_category)
        self.backbutton.clicked.connect(self.close)

        self.update_category_table()

    def update_edit_button_state(self, text):
        self.editButton.setEnabled(bool(text.strip()))

    def update_category_table(self):
        selected_type = self.typeComboBox.currentText().strip()
        db_type = 'income' if selected_type == "درآمد" else "expense"

        self.cursor.execute("SELECT id, name FROM categories WHERE type = ?", (db_type,))
        rows = self.cursor.fetchall()

        self.categoriesTable.setRowCount(len(rows))
        self.categoriesTable.setColumnCount(2)
        self.categoriesTable.setHorizontalHeaderLabels(["شناسه", "دسته‌بندی"])
        self.categoriesTable.setLayoutDirection(QtCore.Qt.RightToLeft)

        self.categoryComboBox.clear()

        for i, row in enumerate(rows):
            id_, name = row
            self.categoryComboBox.addItem(name)

            id_item = QtWidgets.QTableWidgetItem(str(id_))
            name_item = QtWidgets.QTableWidgetItem(name)

            id_item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
            id_item.setTextAlignment(QtCore.Qt.AlignCenter)
            name_item.setTextAlignment(QtCore.Qt.AlignCenter)

            self.categoriesTable.setItem(i, 0, id_item)
            self.categoriesTable.setItem(i, 1, name_item)

    def add_category(self):
        name = self.categoryComboBox.currentText().strip()
        if not name:
            warning_text = (
                "<b style='color:#c00;'>خطا:</b><br>"
                "<span style='color:#333;'>نام <b>دسته‌بندی</b> نمی‌تواند خالی باشد </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطای ورودی")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        selected_type = self.typeComboBox.currentText().strip()
        db_type = 'income' if selected_type == "درآمد" else "expense"

        self.cursor.execute("SELECT id FROM categories WHERE name = ? AND type = ?", (name, db_type))
        if self.cursor.fetchone():
            info_text = (
                "<b style='color:#f57c00;'>اطلاع:</b><br>"
                "<span style='color:#333;'>این <b>دسته‌بندی</b> قبلاً ثبت شده است </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Information)
            msgbox.setWindowTitle("اطلاع‌رسانی")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(info_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        self.cursor.execute("INSERT INTO categories (name, type) VALUES (?, ?)", (name, db_type))
        self.conn.commit()
        self.update_category_table()
        success_text = (
            "<b style='color:green;'>موفقیت:</b><br>"
            "<span style='color:#333;'>دسته‌بندی با موفقیت اضافه شد </span>"
        )

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Information)
        msgbox.setWindowTitle("ثبت موفق")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(success_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msgbox.exec()

    def edit_category(self):
        row = self.categoriesTable.currentRow()
        if row == -1:
            warning_text = (
                "<b style='color:#f57c00;'>هشدار:</b><br>"
                "<span style='color:#333;'>لطفاً یک <b>دسته</b> را از جدول انتخاب کنید </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("هشدار")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        category_id_item = self.categoriesTable.item(row, 0)
        name_item = self.categoriesTable.item(row, 1)

        if not category_id_item or not name_item:
            warning_text = (
                "<b style='color:#d32f2f;'>خطا:</b><br>"
                "<span style='color:#333;'>اطلاعات <b>دسته‌بندی</b> ناقص است. لطفاً همه فیلدها را کامل کنید </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        category_id = category_id_item.text().strip()
        new_name = name_item.text().strip()

        if not new_name:
            warning_text = (
                "<b style='color:#c62828;'>خطا:</b><br>"
                "<span style='color:#444;'>نام <b>دسته‌بندی</b> نمی‌تواند <u>خالی</u> باشد. لطفاً یک نام معتبر وارد کنید </span>"
            )

            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("خطا")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        self.cursor.execute("UPDATE categories SET name = ? WHERE id = ?", (new_name, category_id))
        self.conn.commit()
        success_text = (
            "<b style='color:#2e7d32;'>موفقیت </b><br>"
            "<span style='color:#444;'>تغییرات <b>دسته‌بندی</b> با موفقیت ذخیره شد.</span>"
        )

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Information)
        msgbox.setWindowTitle("موفقیت")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(success_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msgbox.exec()
        self.update_category_table()

    def remove_category(self):
        row = self.categoriesTable.currentRow()
        if row == -1:
            warning_text = (
                "<b style='color:#c00;'>هشدار:</b><br>"
                "لطفاً یک دسته‌بندی را انتخاب کنید<br><br>"
            )
            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Warning)
            msgbox.setWindowTitle("حذف دسته‌بندی و تراکنش‌ها")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(warning_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.setDefaultButton(QtWidgets.QMessageBox.Ok)
            msgbox.exec()
            return

        category_id = self.categoriesTable.item(row, 0).text()

        warning_text = (
            "<b style='color:#c00;'>هشدار:</b><br>"
            "با حذف این دسته‌بندی، تمام تراکنش‌های مرتبط نیز حذف خواهند شد.<br><br>"
            "<b>آیا از ادامه‌ی عملیات مطمئن هستید؟</b>"
        )

        msgbox = QtWidgets.QMessageBox()
        msgbox.setIcon(QtWidgets.QMessageBox.Warning)
        msgbox.setWindowTitle("حذف دسته‌بندی و تراکنش‌ها")
        msgbox.setTextFormat(QtCore.Qt.RichText)
        msgbox.setText(warning_text)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        msgbox.setDefaultButton(QtWidgets.QMessageBox.No)
        confirm = msgbox.exec()

        if confirm == QtWidgets.QMessageBox.Yes:
            from dbfunctions import remove_category
            remove_category(category_id)
            self.update_category_table()
            
            success_text = (
                "<b'>موفقیت:</b><br>"
                "دسته‌بندی و تراکنش‌های مرتبط با موفقیت حذف شدند.<br><br>"
            )
            
            msgbox = QtWidgets.QMessageBox()
            msgbox.setIcon(QtWidgets.QMessageBox.Information)
            msgbox.setWindowTitle("نتیجه عملیات")
            msgbox.setTextFormat(QtCore.Qt.RichText)
            msgbox.setText(success_text)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msgbox.setDefaultButton(QtWidgets.QMessageBox.Ok)
            msgbox.setStyleSheet("""
                QMessageBox {
                    color: white;
                    font-family: 'Tahoma';
                    font-size: 13px;
                }
            """)
            msgbox.exec()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    global window1, window2, window3, window4, window5
    window1 = Main()
    window2 = SignInPage()
    window3 = SignUpPage()
    window4 = OtpPage()
    window1.show()
    sys.exit(app.exec())
    